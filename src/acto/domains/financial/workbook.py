"""Bounded workbook presentation for the synthetic financial example."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def workbook_grid(path: Path, *, formulas: bool) -> dict[str, Any]:
    """Return the review rows without executing workbook content."""

    workbook = load_workbook(path, data_only=not formulas, read_only=True)
    sheet = workbook["Forecast"]
    columns = ["Metric", "FY25A", "FY26E", "FY27E", "FY28E", "FY29E"]
    source_columns = [1, 4, 5, 6, 7, 8]
    row_numbers = [5, 8, 12, 13, 14, 16, 17]
    rows: list[list[Any]] = []

    for row_number in row_numbers:
        row: list[Any] = []
        for column_number in source_columns:
            value = sheet.cell(row=row_number, column=column_number).value
            if isinstance(value, float):
                value = round(value, 3)
            row.append(value)
        rows.append(row)

    return {"columns": columns, "rows": rows}
