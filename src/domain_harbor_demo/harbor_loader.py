from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
import json

from openpyxl import load_workbook


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _reward(result: dict[str, Any]) -> float | None:
    rewards = (result.get("verifier_result") or {}).get("rewards") or {}
    if "reward" in rewards:
        return float(rewards["reward"])
    if rewards:
        return float(next(iter(rewards.values())))
    return None


@dataclass(frozen=True)
class TrialRecord:
    directory: Path
    result: dict[str, Any]
    trajectory: dict[str, Any]
    manifest: list[dict[str, Any]]
    context: dict[str, Any]

    @property
    def id(self) -> str:
        return str(self.result["id"])

    @property
    def name(self) -> str:
        return str(self.result["trial_name"])

    @property
    def task_name(self) -> str:
        return str(self.result["task_name"])

    @property
    def agent_name(self) -> str:
        return str((self.result.get("agent_info") or {}).get("name", "unknown"))

    @property
    def model_name(self) -> str:
        model = ((self.result.get("agent_info") or {}).get("model_info") or {}).get(
            "name"
        )
        return str(model or "unknown")

    @property
    def reward(self) -> float | None:
        return _reward(self.result)

    @property
    def artifact_path(self) -> Path | None:
        for entry in self.manifest:
            destination = entry.get("destination", "")
            if entry.get("status") == "ok" and destination.lower().endswith(".xlsx"):
                path = self.directory / destination
                if path.exists():
                    return path
        return None

    @property
    def status(self) -> str:
        return "Errored" if self.result.get("exception_info") else "Completed"


@dataclass(frozen=True)
class HarborJob:
    directory: Path
    result: dict[str, Any]
    trials: tuple[TrialRecord, ...]

    @property
    def id(self) -> str:
        return str(self.result["id"])

    @property
    def pass_count(self) -> int:
        return sum(1 for trial in self.trials if trial.reward == 1.0)

    @property
    def mean_reward(self) -> float:
        scores = [trial.reward for trial in self.trials if trial.reward is not None]
        return sum(scores) / len(scores) if scores else 0.0


@dataclass(frozen=True)
class LoadedHarborSource:
    """A Harbor job resolved from either a job directory or one of its ATIF files."""

    job: HarborJob
    requested_path: Path
    job_directory: Path
    kind: str
    selected_trial_name: str | None = None


def load_harbor_job(job_dir: str | Path) -> HarborJob:
    directory = Path(job_dir).resolve()
    result = _read_json(directory / "result.json")
    trials: list[TrialRecord] = []

    for trial_result in result.get("trial_results", []):
        trial_dir = directory / trial_result["trial_name"]
        stored_result = _read_json(trial_dir / "result.json")
        trajectory = _read_json(trial_dir / "agent" / "trajectory.json")
        manifest = _read_json(trial_dir / "artifacts" / "manifest.json")
        context_path = trial_dir / "domain_context.json"
        context = _read_json(context_path) if context_path.exists() else {}
        trials.append(
            TrialRecord(
                directory=trial_dir,
                result=stored_result,
                trajectory=trajectory,
                manifest=manifest,
                context=context,
            )
        )

    return HarborJob(directory=directory, result=result, trials=tuple(trials))


def load_harbor_source(source: str | Path) -> LoadedHarborSource:
    """Resolve a Harbor job directory, trial directory, or ATIF trajectory path.

    A standalone ATIF file does not contain verifier results or collected artifacts,
    so the expert workspace requires it to live inside a Harbor trial directory.
    """

    requested_path = Path(source).expanduser().resolve()
    if not requested_path.exists():
        raise FileNotFoundError(f"Harbor source does not exist: {requested_path}")

    job_directory: Path | None = None
    selected_trial_name: str | None = None
    kind = "Harbor job"

    if requested_path.is_file():
        if requested_path.suffix.lower() != ".json":
            raise ValueError("Choose a Harbor job directory or a JSON ATIF trajectory.")

        payload = _read_json(requested_path)
        if str(payload.get("schema_version", "")).startswith("ATIF-"):
            if requested_path.parent.name != "agent":
                raise ValueError(
                    "This is an ATIF file, but it is not inside "
                    "<job>/<trial>/agent/. Choose the enclosing Harbor job or its "
                    "agent/trajectory.json so rewards and artifacts can be reviewed."
                )
            trial_directory = requested_path.parent.parent
            job_directory = trial_directory.parent
            selected_trial_name = trial_directory.name
            kind = "Harbor ATIF"
        elif "trial_results" in payload:
            job_directory = requested_path.parent
        else:
            raise ValueError(
                "The JSON file is neither a Harbor job result nor an ATIF trajectory."
            )
    else:
        directory_result = requested_path / "result.json"
        if directory_result.exists():
            payload = _read_json(directory_result)
            if "trial_results" in payload:
                job_directory = requested_path
            elif "trial_name" in payload:
                job_directory = requested_path.parent
                selected_trial_name = str(payload["trial_name"])
                kind = "Harbor trial"
        elif requested_path.name == "agent" and (
            requested_path / "trajectory.json"
        ).exists():
            job_directory = requested_path.parent.parent
            selected_trial_name = requested_path.parent.name
            kind = "Harbor ATIF"

    if job_directory is None or not (job_directory / "result.json").exists():
        raise ValueError(
            "Could not find the enclosing Harbor job. Choose the job directory or "
            "<job>/<trial>/agent/trajectory.json."
        )

    job = load_harbor_job(job_directory)
    if selected_trial_name and not any(
        trial.name == selected_trial_name for trial in job.trials
    ):
        raise ValueError(
            f"The selected ATIF trial '{selected_trial_name}' is not listed in the "
            "enclosing Harbor job result."
        )

    return LoadedHarborSource(
        job=job,
        requested_path=requested_path,
        job_directory=job_directory,
        kind=kind,
        selected_trial_name=selected_trial_name,
    )


def workbook_grid(path: Path, *, formulas: bool) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=not formulas, read_only=True)
    sheet = workbook["Forecast"]
    columns = ["Metric", "FY25A", "FY26E", "FY27E", "FY28E", "FY29E"]
    source_columns = [1, 4, 5, 6, 7, 8]
    row_numbers = [5, 8, 12, 13, 14, 16, 17]
    rows: list[list[Any]] = []

    for row_number in row_numbers:
        row: list[Any] = []
        for column_number in source_columns:
            value = sheet.cell(row=row_number, column=column_number).value
            if isinstance(value, float):
                value = round(value, 3)
            row.append(value)
        rows.append(row)

    return {"columns": columns, "rows": rows}


def trajectory_rows(trajectory: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for step in trajectory.get("steps", []):
        tool_calls = step.get("tool_calls") or []
        action = ""
        if tool_calls:
            call = tool_calls[0]
            action = str(call.get("function_name", "tool"))
            arguments = call.get("arguments") or {}
            command = arguments.get("command") or arguments.get("path") or ""
            if command:
                action = f"{action}: {command}"

        observation = step.get("observation") or {}
        observation_results = observation.get("results") or []
        outcome = ""
        if observation_results:
            outcome = str(observation_results[0].get("content", ""))

        rows.append(
            {
                "step": str(step.get("step_id", "")),
                "source": str(step.get("source", "")),
                "message": str(step.get("message", "") or ""),
                "action": action,
                "outcome": outcome,
            }
        )
    return rows


def trajectory_metrics(trajectory: dict[str, Any]) -> dict[str, int | float]:
    """Derive comparable behavior and efficiency metrics from an ATIF trajectory."""

    steps = trajectory.get("steps") or []
    tool_calls = [
        tool_call
        for step in steps
        for tool_call in (step.get("tool_calls") or [])
    ]
    observation_count = sum(
        len(((step.get("observation") or {}).get("results") or []))
        for step in steps
    )

    validation_terms = (
        "check",
        "inspect",
        "open",
        "read",
        "recalculate",
        "test",
        "validate",
        "verify",
    )
    change_terms = ("=", "edit", "patch", "replace", "set ", "update", "write")
    validation_actions = 0
    change_actions = 0
    for tool_call in tool_calls:
        arguments = json.dumps(tool_call.get("arguments") or {}).lower()
        validation_actions += int(any(term in arguments for term in validation_terms))
        change_actions += int(any(term in arguments for term in change_terms))

    elapsed_seconds = 0.0
    timestamps = [step.get("timestamp") for step in steps if step.get("timestamp")]
    if len(timestamps) >= 2:
        try:
            started = datetime.fromisoformat(str(timestamps[0]).replace("Z", "+00:00"))
            finished = datetime.fromisoformat(str(timestamps[-1]).replace("Z", "+00:00"))
            elapsed_seconds = max(0.0, (finished - started).total_seconds())
        except ValueError:
            elapsed_seconds = 0.0

    final_metrics = trajectory.get("final_metrics") or {}
    prompt_tokens = int(final_metrics.get("total_prompt_tokens") or 0)
    completion_tokens = int(final_metrics.get("total_completion_tokens") or 0)
    cached_tokens = int(final_metrics.get("total_cached_tokens") or 0)
    cost_usd = float(final_metrics.get("total_cost_usd") or 0.0)

    return {
        "steps": len(steps),
        "tool_calls": len(tool_calls),
        "observations": observation_count,
        "validation_actions": validation_actions,
        "change_actions": change_actions,
        "elapsed_seconds": elapsed_seconds,
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "cached_tokens": cached_tokens,
        "total_tokens": prompt_tokens + completion_tokens,
        "cost_usd": cost_usd,
        "cost_per_action": cost_usd / len(tool_calls) if tool_calls else 0.0,
    }
